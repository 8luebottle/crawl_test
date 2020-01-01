# Happy New Year
# Jan 1, 2020
# IMDB's Top Box Office

import requests

from bs4      import BeautifulSoup as bs
from openpyxl import Workbook

# Set Up
url     = "https://www.imdb.com/chart/boxoffice/?ref_=nv_ch_cht" 
content = requests.get(url)
soup    = bs(content.text, 'html.parser')
top_ten = 10 # Total Movie Count

# Make Excel files
wb = Workbook()
ws = wb.create_sheet('Top Ten Box Office',0)

def top_ten_box_office():
    # Titles
    title_list = soup.find_all('td', {'class':'titleColumn'})  
    title_box = list()
    for titles in title_list:
        for title in titles.find_all('a'):
            title_box.append(title.text)

    # Gross
    gross_list = soup.find_all('td', {'class':'ratingColumn'})
    gross_box  = list()
    for grosses in gross_list:
        for gross in grosses.find_all('span'):
            gross_box.append(gross.text)

    # Weeks
    week_list = soup.find_all('td', {'class':'weeksColumn'})
    week_box = list()
    for weeks in week_list:
        for week in weeks:
            week_box.append(week)

    def table_maker(counts, title, gross, week):
        for i in range(counts):
            ws.cell(i+1,1, title[i])
            ws.cell(i+1,2, gross[i])
            ws.cell(i+1,3, week[i])

    return table_maker(top_ten, title_box, gross_box, week_box)

top_ten_box_office()
wb.save('top_ten.xlsx')
